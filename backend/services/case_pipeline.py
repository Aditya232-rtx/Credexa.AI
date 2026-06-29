from __future__ import annotations

import os
import uuid
import tempfile
import concurrent.futures
from pathlib import Path
from typing import Any, Dict, List, Sequence

from anomaly.main import detect_anomalies
from consistency.cross_doc import cross_check_documents, extract_entities
from forensics.file_forensics import inspect_office_file, inspect_pdf, inspect_pdf_fonts
from forensics.math_validator import validate_bank_statement_xlsx, validate_financials
from forensics.prnu_check import analyze_prnu
from forensics.visual_forensics import run_ela, run_dct_ghost
from ingestion.loader import read_document
from router.classifier import DocumentRouter
from scoring.main import score_case
from utils.encryption import decrypt_data
from db.connection import get_db_connection

LAYER_TIMEOUT = int(os.environ.get("LAYER_TIMEOUT", "120"))  # seconds per document


class CasePipeline:
    def __init__(self, base_dir: Path, db_path: str, documenttypes_path: Path):
        self.base_dir = base_dir
        self.db_dsn = db_path  # We repurpose db_path as the DSN string
        self.documenttypes_path = documenttypes_path
        self.router = DocumentRouter(documenttypes_path=documenttypes_path)

    def _connect(self):
        return get_db_connection()

    def _insert_flag(self, cur, case_id: str, document_id: str | None, flag: Dict[str, Any]) -> None:
        cur.execute(
            "INSERT INTO flags (id, case_id, document_id, layer, finding, severity, score) VALUES (%s, %s, %s, %s, %s, %s, %s)",
            (
                str(uuid.uuid4()),
                case_id,
                document_id,
                flag.get("layer", "Analysis"),
                flag.get("finding", ""),
                flag.get("severity", "low"),
                int(flag.get("score", 0)),
            ),
        )

    def _process_single_document(self, case_id: str, document: Dict[str, Any]) -> Dict[str, Any] | None:
        encrypted_file_path = self.base_dir / "uploads" / case_id / f"{document['id']}.enc"
        if not encrypted_file_path.exists():
            return None

        file_path = None
        temp_file = None
        try:
            raw_ft = (document["file_type"] or "").lower()
            if "/" in raw_ft:
                mime_ext_map = {
                    "image/jpeg": ".jpg", "image/jpg": ".jpg", "image/png": ".png",
                    "image/tiff": ".tif", "image/bmp": ".bmp",
                    "application/pdf": ".pdf",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
                    "application/vnd.ms-excel": ".xls",
                }
                file_ext = mime_ext_map.get(raw_ft, "." + raw_ft.split("/")[-1])
            elif raw_ft.startswith("."):
                file_ext = raw_ft
            elif raw_ft:
                file_ext = "." + raw_ft
            else:
                file_ext = Path(document.get("file_name", "")).suffix or ".bin"

            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=file_ext)
            try:
                encrypted_content = encrypted_file_path.read_bytes()
                temp_file.write(decrypt_data(encrypted_content))
                temp_file.flush()
                file_path = Path(temp_file.name)
            finally:
                temp_file.close()

            payload = read_document(file_path)
            payload["document_id"] = document["id"]
            payload["file_size"] = int(document["file_size"] or 0)

            category = self.router.classify_document(payload)
            text = payload.get("text", "") or ""
            tables = payload.get("tables", []) or []
            metadata = payload.get("metadata", {}) or {}
            pages = payload.get("pages", []) or []

            flags: List[Dict[str, Any]] = []
            is_xlsx_or_csv = file_ext in (".xlsx", ".csv")
            if file_ext == ".pdf":
                flags.extend(inspect_pdf(str(file_path), metadata))
                flags.extend(inspect_pdf_fonts(str(file_path)))
            elif file_ext in (".docx", ".xlsx", ".xls", ".csv"):
                flags.extend(inspect_office_file(str(file_path), metadata))

            if pages:
                for page in pages:
                    image_path = page.get("image_path")
                    if image_path and os.path.exists(image_path):
                        flags.extend(run_ela(str(image_path)))
                        flags.extend(run_dct_ghost(str(image_path)))
                        flags.extend(analyze_prnu(str(image_path)))
                        try:
                            os.remove(image_path)
                        except Exception:
                            pass

            if is_xlsx_or_csv:
                try:
                    import openpyxl

                    workbook = openpyxl.load_workbook(file_path, data_only=False)
                    workbook_data = openpyxl.load_workbook(file_path, data_only=True)
                    sheets_data: Dict[str, Any] = {}
                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        data_sheet = workbook_data[sheet_name]
                        sheet_rows: List[List[Dict[str, Any]]] = []
                        for row_index, row in enumerate(sheet.iter_rows()):
                            row_data: List[Dict[str, Any]] = []
                            for col_index, cell in enumerate(row):
                                value_cell = data_sheet.cell(row=row_index + 1, column=col_index + 1)
                                if cell.value is None and value_cell.value is None:
                                    continue
                                row_data.append(
                                    {
                                        "coordinate": cell.coordinate,
                                        "formula": cell.value if cell.data_type == "f" else None,
                                        "value": value_cell.value,
                                        "hyperlink": cell.hyperlink.target if cell.hyperlink else None,
                                    }
                                )
                            if row_data:
                                sheet_rows.append(row_data)
                        sheets_data[sheet_name] = {"hidden": sheet.sheet_state == "hidden", "rows": sheet_rows}
                    flags.extend(validate_bank_statement_xlsx(sheets_data))
                except Exception:
                    pass
            else:
                if any(token in text.lower() for token in ["balance sheet", "profit", "loss", "reconciliation", "bank statement"]):
                    flags.extend(validate_financials(text, tables, file_ext))

            entities = extract_entities(text)
            return {
                "document_result": {
                    "document_id": document["id"],
                    "file_name": document["file_name"],
                    "category": category,
                    "flags": flags,
                    "entities": entities,
                    "text_preview": text[:500],
                },
                "doc_entity": {"doc_id": document["id"], "entities": entities},
                "flags": flags,
            }
        finally:
            if file_path and file_path.exists():
                file_path.unlink()

    def process_case(self, case_id: str) -> Dict[str, Any]:
        with self._connect() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT * FROM cases WHERE id = %s", (case_id,))
                case = cur.fetchone()
                if not case:
                    raise ValueError(f"Case not found: {case_id}")

                cur.execute("SELECT * FROM documents WHERE case_id = %s ORDER BY file_name", (case_id,))
                documents = cur.fetchall()
                document_results: List[Dict[str, Any]] = []
                doc_entities: List[Dict[str, Any]] = []
                all_flags: List[Dict[str, Any]] = []

                cur.execute("UPDATE cases SET status = %s, updated_at = CURRENT_TIMESTAMP WHERE id = %s", ("processing", case_id))
                cur.execute("DELETE FROM flags WHERE case_id = %s", (case_id,))

                for document in documents:
                    try:
                        with concurrent.futures.ThreadPoolExecutor(max_workers=1) as executor:
                            future = executor.submit(self._process_single_document, case_id, document)
                            doc_result = future.result(timeout=LAYER_TIMEOUT)
                    except concurrent.futures.TimeoutError:
                        logger.warning(f"Document {document['id']} processing timed out after {LAYER_TIMEOUT}s")
                        continue
                    except Exception as e:
                        logger.warning(f"Document {document['id']} processing failed: {e}")
                        continue

                    if doc_result is None:
                        continue
                    doc_results = doc_result
                    document_results.append(doc_results["document_result"])
                    doc_entities.append(doc_results["doc_entity"])
                    for flag in doc_results["flags"]:
                        self._insert_flag(cur, case_id, document["id"], flag)
                    all_flags.extend(doc_results["flags"])

                cross_doc_flags = cross_check_documents(doc_entities)
                for flag in cross_doc_flags:
                    self._insert_flag(cur, case_id, None, flag)
                all_flags.extend(cross_doc_flags)

                anomaly_result = detect_anomalies(document_results)
                for flag in anomaly_result.flags:
                    self._insert_flag(cur, case_id, None, flag)
                all_flags.extend(anomaly_result.flags)

                scoring = score_case(all_flags, anomaly_result.score)
                cur.execute(
                    "UPDATE cases SET status = %s, risk_score = %s, updated_at = CURRENT_TIMESTAMP WHERE id = %s",
                    (scoring["status"], scoring["risk_score"], case_id),
                )
                cur.execute(
                    "INSERT INTO audit_log (id, case_id, action, details) VALUES (%s, %s, %s, %s)",
                    (str(uuid.uuid4()), case_id, "analysis_completed", scoring["explanation"]),
                )
            conn.commit()

            return {
                "case_id": case_id,
                "risk_score": scoring["risk_score"],
                "status": scoring["status"],
                "explanation": scoring["explanation"],
                "documents": document_results,
                "flags": all_flags,
            }